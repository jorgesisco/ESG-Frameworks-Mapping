from cmath import nan
from distutils.filelist import findall
import pandas as pd
import numpy as np
import re

import pdfplumber
from tabula import read_pdf
import openpyxl
import camelot.io as camelot


class ExtractPDFTables:

	# Initializing contructor variables
	def __init__(self, file_path, page_range=False, area=False, flavor=None, strip_text=None):
		self.file_path = file_path #string with file path
		self.page_range = page_range # defined page range with tables
		self.area = area # Table area, libraris like tabula allows to specify table area
		
		#Camelot parameters
		self.flavor = flavor
		self.strip_text = strip_text
	
	#Table Extrator with Camelot Library
	def getTablesCamelot(self, flavor=None, strip_text=None):
		tables = camelot.read_pdf(self.file_path, 
								  pages=f'{self.page_range[0]}-{self.page_range[1]}', 
								  flavor=f'{self.flavor}', 
								  strip_text=f'{self.strip_text}')

		frames = []
		for i in tables:
			try:
				table = i.df
				frames.append(table)
			except:
				pass

		df =  pd.concat(frames)
		return df

	def getTablesPDFplumber(self):
		pdf = pdfplumber.open(self.file_path, pages=self.page_range)
		frames = []

		for i in range(0, len(self.page_range)):
			try:
				page = pdf.pages[i]
				table = page.extract_table()
				frames.append(pd.DataFrame(table))
			
			except:
				pass

		df =  pd.concat(frames)
		df = df.drop_duplicates()

		return df



	def getCode(self, df,
				column,
				regex, delimiter, 
				code_column='Code',
				remove_char=None):

			df[code_column] = df[column].str.findall(regex).apply(set).str.join(delimiter)

			if remove_char != None:
				df[code_column] = df[code_column].str.replace(remove_char, '', regex=False)

			return df

    # This metho will extract strings from designated column and add them to a new column
	def separate_code(self, df, column, delimiter):
		
		df = df.assign(GRI_Code=df[f'{column}'].str.split(delimiter)).explode(f'{column}')
	
		return df

	#Merge codes that has similar code from other framework (this is good for mapping in the framework that is the code in common with the rest)
	def merge_codes(self, df, ref_col, code_column):
		array_agg = lambda x: '\n'.join(x.astype(str))
		df = df.groupby([ref_col], as_index=False).agg({code_column: array_agg})
		return df



	# Getting tables from PDF SDG Linked to GRI
	def getTablesSDG_GRI(self, sdg=None):

		if sdg != None:
					sdg_df = pd.read_csv(sdg)
		
		'''This method requires to add a nested list with 2 page ranges,
		   with this approach, we avoid scanning a non table on page 73
		'''
		if any(isinstance(i, list) for i in self.page_range) == False:
			return print("WARNING: page_range should be a nested list\n eg: page_range = [list(range(3, 73)), list(range(74, 99))]\n Why? the SDG-GRI pdf file has a page in the middle that is\n not a table, getTablesSDG_GRI method iterates the first\n range of pages an then the other one.")

		else:
			# Gets all the tables from first page range
			pdf = read_pdf(self.file_path, stream=True, pages = self.page_range[0],
						   area = self.area, multiple_tables=False)

			#renaming "sources" column to "source" (same column name in the reamining tables)
			pdf[0].rename(columns = {'Sources':'Source'}, inplace = True)
			
			# Getting the remaining tables
			pdf_ = read_pdf(self.file_path, stream=True, pages = self.page_range[1],
							area = self.area, multiple_tables=False )

			#Merging all the tables in one pandas dataframe
			pdf[0] = pdf[0].append(pdf_[0])

			# Renaming variable name
			df = pdf[0]
			df = df[df.Target != 'Target']

			
			array_agg = lambda x: '\n'.join(x.astype(str))

			df = df.fillna(method='ffill')
			df = df.groupby(['Target', 'Disclosure'], as_index=False).agg({'Available Business Disclosures': array_agg})

			structuringApproach = input('Are you going to map GRI on SDG sheet? (yes or no): ')

			if structuringApproach == 'yes':
				print('Structured GRI for SDG sheet, this data can be mapped on SDG excel sheet with the GRI Disclosures')

				df = df.groupby(['Target'], as_index=False).agg({'Available Business Disclosures': array_agg, 'Disclosure':array_agg})
				# df = df.drop(labels = 'Target',axis = 1).groupby(df['Target'].mask(df['Target']==' ').ffill()).agg(', '.join).reset_index()

				df.rename(columns = {'Target':'SDG_Target', 'Available Business Disclosures': 'GRI Available Business Disclosures', 
									 'Disclosure': 'GRI_Disclosure'}, inplace = True)
			
				if sdg != None:
					df = pd.merge(df, sdg_df, on=['SDG_Target'], how='right')
					df = df[['SDG_Target', 'SDG Description','GRI_Disclosure', 'GRI Available Business Disclosures']]
				
				# df = df.dropna()

			elif structuringApproach == 'no':

				df.rename(columns = {'Target':'SDG_Target', 'Available Business Disclosures': 'GRI Available Business Disclosures', 
									 'Disclosure': 'GRI_Disclosure'}, inplace = True)
				
				if sdg != None:
					df = pd.merge(df, sdg_df, on=['SDG_Target'], how='right')
					df = df[['SDG_Target', 'SDG Description','GRI_Disclosure', 'GRI Available Business Disclosures']]
				

			else:
				print("WARNING: please run the script again and choose 'yes' or 'no' ") 

			return df


	# Getting tables from PDF GRI Linked to COH4B
	def getTablesGRI_COH4B(self):
		
		pdf = pdfplumber.open(self.file_path, pages=self.page_range)
		frames = []

		for i in range(0, len(self.page_range)):
			try:
				page = pdf.pages[i]
				table = page.extract_table()
				frames.append(pd.DataFrame(table))
			
			except:
				pass

		df =  pd.concat(frames)
		df = df.drop_duplicates()

		return df
	
	# Getting tables from PDF TCFD Linked to GRI
	def getTablesTCFD_GRI(self):
		# pdf = read_pdf(self.file_path, pages=self.page_range, area = self.area, multiple_tables=True)
		pdf = pdfplumber.open(self.file_path, pages=self.page_range)

		frames = []

		for i in range(0, len(self.page_range)):

			page = pdf.pages[i]
			table = page.debug_tablefinder()
			req_table = table.tables[0] 

			cells = req_table.cells

			for cell in cells[0:len(cells)]:
			    page.crop(cell).extract_words()
			
			data = page.extract_table()
			df = pd.DataFrame(data)
			df = pd.DataFrame(data[2:],columns=data[1])
			df.iloc[:, 0] = df.iloc[:, 0].str.replace('\n' , ' ')
			# df.iloc[:, 0] = df.iloc[:, 0].str.replace('Governance' , '')
			df.iloc[:, 1] = df.iloc[:, 1].str.replace('\n' , ' ')
			frames.append(df)


		df =  pd.concat(frames)
		df = df.drop_duplicates()

		return df

	# Getting tables from PDF CASS CSR 4.0 - GRI 2016

	def getTablesCASS_GRI(self):
		pdf = pdfplumber.open(self.file_path, pages=self.page_range)
		frames = []

		for i in range(0, len(self.page_range)):
			try:
				page = pdf.pages[i]
				table = page.extract_table()
				frames.append(pd.DataFrame(table))
				
			except:
				pass

		df =  pd.concat(frames)
		df.columns = df.iloc[0]
		df = df[1:]
		# find_duplicate = df['KPI \n(CASSCSR-4.0)'].isin(['KPI \n(CASSCSR-4.0)'])

		# df = df[~find_duplicate]
		# Inserting above values on NaN (None) cells
		df = df.fillna(method='ffill')
		df.columns.values[0] = "ID_CASS_CSR"
		df.columns.values[2] = "ID_GRI"



		return df

	def getTablesCDP_GRI(self):
		# regex = r'(\d\d\d-\d\d)'
		regex = r'\d\d\d-\d\d*| \d\d\d-\d\d*-\w'
		
		pdf = pdfplumber.open(self.file_path, pages=self.page_range)
		frames = []

		for i in range(0, len(self.page_range)):
			try:
				page = pdf.pages[i]
				table = page.extract_table()
				frames.append(pd.DataFrame(table))
					
			except:
				pass

		df =  pd.concat(frames)
		df.columns = df.iloc[0]
		df = df[1:]

		structuringApproach = input("Are you trying to map CDP Climate Change 2017 - GRI 2016 press '1'  or CDP Water 2018 - GRI?? press '2'")
		

		if structuringApproach == str(1):
			structuringApproach_2 = input('Do you want to map CDP on GRI sheet (yes or no)? - if no, then the dataframe will be optimal for mapping GRI on CDP sheet:')

			if structuringApproach_2 == 'yes':
				df['GRI ID'] = df['GRI'].str.findall(regex).str.join(' ') # will be needed for mapping gri on csp sheet :D
				df['GRI ID'] = df['GRI ID'].str.split()
				df = df.explode('GRI ID')
				return df

			elif structuringApproach_2 == 'no':
				df['GRI ID'] = df['GRI'].str.findall(regex).str.join(',\n')
				return df

			else:
				return print("WARNING: Answer 'yes' or 'no' without any typos")

		elif structuringApproach == str(2):
			structuringApproach_2 = input('Do you want to map CDP on GRI sheet (yes or no)? - if no, then the dataframe will be optimal for mapping GRI on CDP sheet:')

			if structuringApproach_2 == 'yes':
				df['GRI ID'] = df['GRI Disclosures'].str.findall(regex).str.join(' ') # will be needed for mapping gri on csp sheet :D
				df['GRI ID'] = df['GRI ID'].str.split()
				df = df.explode('GRI ID')
				return df

			elif structuringApproach_2 == 'no':
				df['GRI ID'] = df['GRI Disclosures'].str.findall(regex).str.join(',\n')
				return df

			else:
				return print("WARNING: Answer 'yes' or 'no' without any typos")
				

	def getTablesCDP_TCFD(self):

		tables = camelot.read_pdf(self.file_path, 
								  pages=f'{self.page_range[0]}-{self.page_range[1]}', 
								  flavor='lattice')

		frames = []
		for i in tables:
			try:
				table = i.df
				frames.append(table)
			except:
				pass

		df =  pd.concat(frames)
		df.columns = df.iloc[0]
		df = df[1:]

		return df

	def getTablesGRI_GRI_OILGAS_OR_COAL(self):
		pdf = pdfplumber.open(self.file_path, pages=self.page_range)
		frames = []

		for i in range(0, len(self.page_range)):
			try:
				page = pdf.pages[i]
				table = page.extract_table()
				frames.append(pd.DataFrame(table))
					
			except:
				pass

		df =  pd.concat(frames)
		df = df.drop_duplicates(keep='first')
		df.columns = df.iloc[0]
		df = df[1:]
		df = df.fillna(method='ffill') # Filling rows below with same value as above 
		df = df.replace(r'\n','', regex=True)

		year = input("with GRI sheet would you like to map? 2016 or 2021?")

		if year == '2016':
			df = df[df['STANDARD'].str.contains(year)]
			# df = df.replace(f' {year}','', regex=True)
			df['GRI Code'] = df['DISCLOSURE'].str.findall(r'\d+-\d+').str.join(' ')
			return df
		
		elif year == '2021':
			df = df[df['STANDARD'].str.contains(year)]
			# df = df.replace(f' {year}','', regex=True)
			df['GRI Code'] = df['DISCLOSURE'].str.findall(r'\d-\d').str.join(' ')
			return df
		
		else:
			return print("Type 2016 or 2017, there are not other options...")

	def getTablesGRI_HKEX20(self):
		pdf = pdfplumber.open(self.file_path, pages=self.page_range)
		frames = []

		for i in range(0, len(self.page_range)):
			try:
				page = pdf.pages[i]
				table = page.extract_table()
				frames.append(pd.DataFrame(table))
					
			except:
				pass

		df =  pd.concat(frames)
		df = df.drop_duplicates(keep='first')
		df.columns = df.iloc[0]
		df = df[1:]
		# note: I GRI16 foundation clauses are not been able to extract from text with regex, mapped them manually in the meantime, still looking for a regex solution

		framework = input('Are you going to map GRI 2016 sheet? yes or no:')

		if framework == 'yes':
			df['GRI_Code'] = df['GRI Standards and Disclosures'].str.findall(r'\d+-\d+').str.join(', ')
			# pandas Explode feature allows to separate GRI code on an unique row with the info to map on excel
			df = df.assign(GRI_Code=df['GRI_Code'].str.split(', ')).explode('GRI_Code')
			return df
		
		elif framework == 'no':
			df['GRI_Code'] = df['GRI Standards and Disclosures'].str.findall(r'\d+-\d+').str.join('\n')
			return df

		return df
	
	def getTablesGRI_SEBI_GRI(self, flavor=None, strip_text=None):
		regex = '\w\d\-\w\d\d\w|\w\d\-\w\d\w|\w\d-\w\d|\w\d\w,\w,\w|\w\d\d\w|\w\d\d|\w\d|^$'

		df = self.getTablesCamelot(flavor, strip_text)
		
		df = df.drop_duplicates(keep='first')
		df.rename(columns = {0:'SEBI - BRSR Framework', 1:'GRI Standards and Disclosures'}, inplace = True)
		df['SEBI - BRSR Framework'] = df['SEBI - BRSR Framework'].replace(' ', '', regex=True)
		df = df[df['SEBI - BRSR Framework'].str.contains(regex)]
		df = df[df["GRI Standards and Disclosures"].str.contains("No direct linkage")==False]
		df = df.drop(labels = 'SEBI - BRSR Framework',axis = 1).astype(str).groupby(df['SEBI - BRSR Framework'].mask(df['SEBI - BRSR Framework']=='').ffill()).agg(' '.join).reset_index()

		# Extract gri 2021 codes
		df['GRI_Disclosures_2021'] = df['GRI Standards and Disclosures'].str.findall('(?<=\s)(\d-\d+)').apply(set).str.join(', ')
		#The funciton with the explode method is to separate each code with all the data for that specific code
		df = df.assign(GRI_Disclosures_2021=df['GRI_Disclosures_2021'].str.split(', ')).explode('GRI_Disclosures_2021')
		# df['GRI Disclosures 2021'] = df['GRI Disclosures 2021'].replace(' ','\n', regex=True)
		df['GRI Standard 2021'] = df['GRI Standards and Disclosures'].str.findall('(GRI \d+: [\w\s]+ 2021)').apply(set).str.join('\n')

		# Extract gri 2016 codes
		df['GRI_Disclosures_2016'] = df['GRI Standards and Disclosures'].str.findall('(\d\d\d-\w+)').apply(set).str.join(', ')
		#The funciton with the explode method is to separate each code with all the data for that specific code
		df = df.assign(GRI_Disclosures_2016=df['GRI_Disclosures_2016'].str.split(', ')).explode('GRI_Disclosures_2016')
		# df['GRI Disclosures 2016'] = df['GRI Disclosures 2016'].replace(' ','\n', regex=True)

		df['GRI Standard 2016'] = df['GRI Standards and Disclosures'].str.findall('(GRI \d+: [\w\s]+ 2016)').apply(set).str.join('\n')
		df['GRI Standard 2018'] = df['GRI Standards and Disclosures'].str.findall('(GRI \d+: [\w\s]+ 2018)').apply(set).str.join('\n')
		df = df[df["GRI Standard 2018"].str.contains("2018")==False]

		return df

	def getTablesGRI_ADX(self):

		df = self.getTablesCamelot(self.flavor, self.strip_text)

		df = df.drop_duplicates(keep='first')
		df.columns = df.iloc[0]
		df = df[1:]
		df= df[['Metric', 'Calculation', 'CorrespondingGRI Standard']]
		df.replace('', np.nan, inplace=True)
		df = df.dropna()

		df = self.getCode(df, code_column='ADX_Code',
                     column='Calculation',
                     regex='\w+.\d+\)', 
                     delimiter='\n', 
                     remove_char=')')


		df = self.getCode(df, code_column='GRI_Code',
							column='CorrespondingGRI Standard',
							regex='\d\d\d:', 
							delimiter='\n', 
							remove_char=':')

		framework = input('Are you mapping on the GRI 2016 sheet? (yes or no):')

		if framework == 'yes':
			df = self.merge_codes(df, 'GRI_Code', 'ADX_Code')
			return df
		
		elif framework == 'no':
			df = df[['Metric', 'Calculation', 'CorrespondingGRI Standard']]
			return df
		
		else:
			print("Enter 'yes' or 'no'")

	def getTablesGRI_BAHRAIN(self):

		df = self.getTablesCamelot()


		return df



	# Funtions needed in some of the extracted dataframes
	def extractDisclosures1(self, df, column, newColumn, regex, method):
		
		values = []

		for i in range(0, len(df[column])):

			try:
				match = method(regex, df[column][i])
				values.append(df[column][i][match.start():match.end()])

			except:
				values.append('No value :(')
				pass
		df[newColumn] = values

		return df

	def extractDisclosures2(self, df, column, newColumn, regex, method_):
		
		values = []

		for i in range(0, len(df[column])):

			try:
				match = method_(regex, df[column].values[i])

				if len(match) > 2:
					values.append(', '.join(match))
				else:
					values.append(''.join(match[0]))

			except:
				values.append('No value :(')
				pass
		df[newColumn] = values

		return df

	def df_gri_tcdf(self, df):

		df_GRI_TCDF = df[['GRI Standards', 'id', 'Recommended \nDisclosures \n(TCFD Framework)']]
		df_GRI_TCDF.rename(columns = {'GRI Standards':'GRI_Standards'}, inplace = True)
		df_GRI_TCDF['GRI_Standards'].str.split(', ')

		df_GRI_TCDF = df_GRI_TCDF.assign(GRI_Standards=df_GRI_TCDF['GRI_Standards'].str.split(', ')).explode('GRI_Standards')

		return df_GRI_TCDF

	'''This methods are usefull in case the gathered table
	   does now shows the columns as headers, but columns
	   are shown as values, we run the method with the
	   gathered dataframe and the row index we want as header. 
	'''
	def setHeaders(self, df, rowIndex):
		headers = df.iloc[rowIndex]
		df = pd.DataFrame(df.values[rowIndex+1:], columns=headers)

		return df

	'''There is a header with None value and the correct value is
	   in the column to the left when choosing that one header 'row 1 header', 
	   the None header as None and the value from column 1 that is suppose to 
	   be an id,
	'''
	def headerSwap(self, df, h1, h2, newh1):

		df.rename(columns = {h1:newh1, h2:h1}, inplace = True)
        
		return df

	''' It is convinient to add a . to the id numbers, to easily match
		with the same values on the excel files and do the mapping.'''
	def addDot(self, df, column):
		df[column] = df[column].str[:-1] + df[column].str[-1] + '.'

		return df

	


class MapLinks2Excel:

	def __init__(self, df, sheet, path_file, 
				first_row=None, excel_ref_column=None, 
				df_ref_column_str=None, df_ref_column_to_add=None,
				excel_column_to_add=None, ESG_to_add=None):

		self.df = df
		self.path_file = path_file
		self.sheet = sheet
		self.first_row = first_row #From which row were ESG refs starts
		self.excel_ref_column = excel_ref_column # column with current esg data used as reference for mapping
		self.df_ref_column_str = df_ref_column_str
		self.df_ref_column_to_add = df_ref_column_to_add
		self.excel_column_to_add = excel_column_to_add
		self.ESG_to_add = ESG_to_add
		
	
	def MapSDG_GRI(self):
		regex = '[+-]?[0-9]+\.-?[0-9a-zA-Z_]+'

		wb = openpyxl.load_workbook(self.path_file)
		ws = wb[self.sheet]

		rows = ws.max_row

		for i in range(1, rows):
			if ws.cell(row=i, column=1).value != None:
				target_cell = ws.cell(row=i, column=1).value
				

				if (re.search(regex, target_cell)):
					target = re.search(regex, target_cell).group()

					try:
						disclosure_to_add = self.df.loc[self.df['SDG_Target'] == target]['GRI_Disclosure'].values
						description_to_add = self.df.loc[self.df['SDG_Target'] == target]['GRI Available Business Disclosures'].values
						
						if str(disclosure_to_add[0]) != 'nan':
							ws.cell(row=i, column=3, value='\n'.join(disclosure_to_add))	
							ws.cell(row=i, column=4, value='\n'.join(description_to_add))
					except:
						pass

		wb.save(self.path_file)

		return f'{self.sheet} sheet from Excel file have bee mapped'
	
	def MapGRI_SDG(self):

		wb = openpyxl.load_workbook(self.path_file)
		ws = wb[self.sheet]
		rows = ws.max_row

		for i in range(1, rows):

			if ws.cell(row=i, column=2).value != None:
				target_cell = ws.cell(row=i+1, column=2).value

				if target_cell != None:
					try:
						if len(self.df[self.df['GRI_Disclosure'] == target_cell]) != 0:
							target_to_add = self.df[self.df.GRI_Disclosure==target_cell].squeeze()['SDG_Target'].values
							# ws.cell(row=i+1, column=6, value='\n '.join(target_to_add))

							disclosure_to_add = self.df[self.df.GRI_Disclosure==target_cell].squeeze()['SDG Description'].values
							# ws.cell(row=i+1, column=7, value='\n '.join(disclosure_to_add))

							if str(target_to_add[0]) != 'nan' and str(disclosure_to_add[0]) != 'nan':
								ws.cell(row=i+1, column=6, value='\n '.join(target_to_add))
								ws.cell(row=i+1, column=7, value='\n '.join(disclosure_to_add))
					except:
						pass

		wb.save(self.path_file)

		return f'{self.sheet} sheet from Excel file have bee mapped'

	def MapCOH4B_GRI(self):
		regex = "[+-]?[0-9]+\."

		wb = openpyxl.load_workbook(self.path_file)
		ws = wb[self.sheet]
		rows = ws.max_row

		''' In case of loading df from csv file, the commented code below
			will fixe and issue with the id column, since the df it's being taking
			from instance tableGRI_COH4B, it is no needed to excute that code.
		'''

		# pdf_tables = pd.read_csv(pdf_tables)
		# pdf_tables['id'] = pdf_tables['id'].astype(str).apply(lambda x: x.replace('.0','.'))

		for i in range(1, rows):
			if ws.cell(row=i, column=1).value == None:
				pass

			else:
				target_cell = ws.cell(row=i, column=1).value

				if (re.search(regex, target_cell)):
					target = re.search(regex, target_cell).group()

					try:
						value_to_add = [self.df.loc[self.df['id'] == target]['C. GRI \nStandards'].values,
										self.df.loc[self.df['id'] == target]['D. GRI disclosures'].values]
						
						ws.cell(row=i, column=3, value=' '.join(value_to_add[0]))
						ws.cell(row=i, column=4, value=' '.join(value_to_add[1]))

					except:
						pass

		wb.save(self.path_file)

		return f'{self.sheet} sheet from Excel file have bee mapped'

	def MapGRI_COH4B(self):

		regex = '^[0-9\.\-\/]+$'

		wb = openpyxl.load_workbook(self.path_file)
		ws = wb[self.sheet]
		rows = ws.max_row

		for i in range(1, rows):
			if ws.cell(row=i, column=2).value == None:
				pass

			else:
				target_cell = ws.cell(row=i, column=2).value
				if(re.search(regex, target_cell)):
					target = re.search(regex, target_cell).group()

					try:
						value_to_add = self.df.loc[self.df['GRI Standards'] == target]['id'].values
						# print(value_to_add)
						ws.cell(row=i, column=10, value=' '.join(value_to_add))

						value_to_add_2 = self.df.loc[self.df['GRI Standards'] == target]['A. COHBP & \ndefinition'].values
						ws.cell(row=i, column=11, value=' '.join(value_to_add_2))
					except:
						pass

		wb.save(self.path_file)

		return f'{self.sheet} sheet from Excel file have bee mapped' 

	def mapTCFD_GRI(self):

		regex = '[a-zA-Z]+\)'
		wb = openpyxl.load_workbook(self.path_file)
		ws = wb[self.sheet]
		rows = ws.max_row

		count = 0

		for i in range(1, rows):
			if ws.cell(row=i, column=2).value != None and count <= len(self.df):
				target_cell = ws.cell(row=i, column=2).value

				if re.search(regex, target_cell):

					target = re.search(regex, target_cell).group()
					target = re.sub('[abc]\)', self.df['id'].values[count], target)

					value_to_add = self.df.loc[self.df['id'] == target]['Related \ncode/\nparagraph'].item()
					value_to_add = re.sub(', ', ' ', value_to_add)
					value_to_add = re.sub('\s\s+', ' ', value_to_add)
					value_to_add = re.sub(' and| with', '', value_to_add)

					value_to_add_2 = self.df.loc[self.df['id'] == target]['Description'].item()

					ws.cell(row=i, column=4, value=value_to_add)

					ws.cell(row=i, column=5, value=value_to_add_2)
					
					count += 1

		wb.save(self.path_file)

	def mapGRI_TCFD(self):

		wb = openpyxl.load_workbook(self.path_file)
		ws = wb[self.sheet]
		rows = ws.max_row

		for i in range(1, rows+1):
			if ws.cell(row=i, column=2).value != None:
				target_cell = ws.cell(row=i, column=2).value

				if target_cell != None:
					value = self.df.loc[self.df['GRI_Standards'] == target_cell]['id'].values
					value_2 =self.df.loc[self.df['GRI_Standards'] == target_cell]['Recommended \nDisclosures \n(TCFD Framework)'].values
					if len(value):
						value_to_add = ' '.join(value)
						ws.cell(row=i, column=8, value=value_to_add)
						# print(value_to_add)

				
					if len(value_2):
						value_to_add_2 = ' \n'.join(value_2)

						ws.cell(row=i, column=9, value=value_to_add_2)

						# print(value_to_add_2)
						# print('--------------------------------')

		

		wb.save(self.path_file)		

	def mapGRI2016_2021(self):
		wb = openpyxl.load_workbook(self.path_file)
		ws = wb[self.sheet]
		rows = ws.max_row

		for i in range(3, rows):
			if ws.cell(row=i, column=2).value == None:
				pass

			else:
				target_cell = ws.cell(row=i, column=2).value
        
        
				if target_cell != None:
					try:	
						
						disclosure_to_add = self.df.loc[self.df['Disclosure Number 2016'] == target_cell]['Disclosure Number 2021'].values
						section_to_add = self.df.loc[self.df['Disclosure Number 2016'] == target_cell]['Section 2021'].values
					
						if disclosure_to_add != False:
							if str(disclosure_to_add[0]) != 'nan':
								ws.cell(row=i, column=4, value=''.join(disclosure_to_add))
								ws.cell(row=i, column=5, value=' '.join(section_to_add))
					except:
						pass

		wb.save(self.path_file)

		print(f"{self.sheet} sheet from Excel file have bee mapped with it's GRI 2021 equivalent")

	def mapGRI2021_2016(self):
		wb = openpyxl.load_workbook(self.path_file)
		ws = wb[self.sheet]
		rows = ws.max_row

		for i in range(3, rows):
			if ws.cell(row=i, column=2).value == None:
				pass

			else:
				target_cell = ws.cell(row=i, column=2).value

        
				if target_cell != None:
					try:
						disclosure_to_add = self.df.loc[self.df['Disclosure Number 2021'] == target_cell]['Disclosure Number 2016'].values
						ws.cell(row=i, column=4, value='\n'.join(disclosure_to_add))
						section_to_add = self.df.loc[self.df['Disclosure Number 2021'] == target_cell]['Section 2016'].values
						ws.cell(row=i, column=5, value=section_to_add[0])	

					except:
						pass

		wb.save(self.path_file)

		print(f"{self.sheet} sheet from Excel file have bee mapped with it's GRI 2021 equivalent")

	def mapCASS_GRI1016(self):
		wb = openpyxl.load_workbook(self.path_file)
		ws = wb[self.sheet]
		rows = ws.max_row
		regex = '\w[+-]?[0-9]+\.+\d+'

		for i in range(3, rows):
			if ws.cell(row=i, column=2).value == None:
				pass

			else:
				target_cell = ws.cell(row=i, column=2).value
				# print(target_cell)

				
				if target_cell != None:
					
					try:
						disclosure_to_add = self.df.loc[self.df['ID_GRI'] == target_cell]['ID_CASS_CSR'].values
						
						
						if len(disclosure_to_add) > 0:
							disclosure_to_add_ = [i[re.search(regex, i).start():re.search(regex, i).end()] for i in disclosure_to_add]
							# print(disclosure_to_add_)
							ws.cell(row=i, column=12, value='\n'.join(disclosure_to_add_))
							
						section_to_add = self.df.loc[self.df['ID_GRI'] == target_cell]['KPI \n(CASSCSR-4.0)'].values
						
						
						if len(section_to_add)>0:
							
							# print(section_to_add)
							ws.cell(row=i, column=13, value='\n'.join(section_to_add))	

					except:
						pass
		wb.save(self.path_file)

		print(f"{self.sheet} sheet from Excel file have bee mapped with it's CASS CSR 4.0 equivalent")

	def mapGRI2016_CASS(self):
		array_agg = lambda x: '\n '.join(x.astype(str))
		array_agg_ = lambda x: ''.join(x.astype(str))
		
		df = self.df.groupby(['ID_CASS_CSR'], 
			 as_index=False).agg({'ID_GRI': array_agg, 
								  'Disclosure title (GRI)': array_agg, 
								  'KPI \n(CASSCSR-4.0)': array_agg_})

		wb = openpyxl.load_workbook(self.path_file)
		ws = wb[self.sheet]

		for i in range(3, len(df['ID_CASS_CSR'])):
			ws.cell(row=i, column=1, value=df['ID_CASS_CSR'][i])
			ws.cell(row=i, column=2, value=df['KPI \n(CASSCSR-4.0)'][i])
			ws.cell(row=i, column=3, value=df['ID_GRI'][i])
			ws.cell(row=i, column=4, value=df['Disclosure title (GRI)'][i])
	
		wb.save(self.path_file)

		print(f"{self.sheet} sheet from Excel file have bee mapped with it's GRI 2016 equivalent") 

	def mapCDPCC17_GRI(self):
		wb = openpyxl.load_workbook(self.path_file)
		ws = wb[self.sheet]
		rows = ws.max_row

		for i in range(3, rows):
			if ws.cell(row=i, column=2).value == None:
				pass

			else:
				target_cell = ws.cell(row=i, column=2).value
		
				if target_cell != None:
					try:
						disclosure_to_add = self.df.loc[self.df['GRI ID'] == target_cell]['CDP'].values
						
						if len(disclosure_to_add) > 0:
							ws.cell(row=i, column=14, value='\n'.join(disclosure_to_add))
					except:
							pass

		wb.save(self.path_file)
		print(f"{self.sheet} sheet from Excel file have bee mapped with it's CDP 2017 equivalent") 

	def mapGRI_CDPCC17(self):
		wb = openpyxl.load_workbook(self.path_file)
		ws = wb[self.sheet]

		regex = '[a-zA-Z]+.+\d'
		r = 3
		for i in self.df['CDP']:
			if re.search(regex, i):
				ws.cell(row=r, column=1, value=i)
				r+=1

		regex_2 = 'GRI \w+\d:'
		r = 3
		for i in self.df['GRI']:
			if i != None and re.search(regex_2, i):
				ws.cell(row=r, column=2, value=i)
				r+=1

				

		wb.save(self.path_file)

	def mapCDP18_GRI(self):
		wb = openpyxl.load_workbook(self.path_file)
		ws = wb[self.sheet]
		rows = ws.max_row

		for i in range(3, rows):
			if ws.cell(row=i, column=2).value == None:
				pass

			else:
				target_cell = ws.cell(row=i, column=2).value

				# print(target_cell)
		
				if target_cell != None:
					try:
						disclosure_to_add = self.df.loc[self.df['GRI ID'] == target_cell]['CDP Water Security Questions'].values
						
						if len(disclosure_to_add) > 0:
							ws.cell(row=i, column=16, value='\n'.join(disclosure_to_add))
						
					except:
							pass

		wb.save(self.path_file)
		print(f"{self.sheet} sheet from Excel file have bee mapped with it's CDP Water 2018 equivalent") 

	def mapGRI_CDP18(self):
		wb = openpyxl.load_workbook(self.path_file)
		ws = wb[self.sheet]

		regex = '[a-zA-Z]+.+\d'
		r = 3
		for i in self.df['CDP Water Security Questions']:
			if re.search(regex, i):
				ws.cell(row=r, column=1, value=i)
				r+=1

		regex_2 = 'GRI \w+\d:'
		r = 3
		for i in self.df['GRI Disclosures']:
			if i != None and re.search(regex_2, i):
				ws.cell(row=r, column=2, value=i)
				r+=1

				

		wb.save(self.path_file)

	def mapEV22(self):
		pass

	def mapCDP_TCFD(self):
		wb = openpyxl.load_workbook(self.path_file)
		ws = wb[self.sheet]

		r=3
		for i in self.df["Question \nnumber (CDP \nclimate change)"]:
			if i != None:
				ws.cell(row=r, column=1, value=i)
				r+=1

		r=3
		for i in self.df['Question text']:
			if i != None:
				ws.cell(row=r, column=2, value=i)
				r+=1

		r=3
		for i in self.df['TCFD recommendations']:
			if i != None:
				ws.cell(row=r, column=3, value=i)
				r+=1
		
		wb.save(self.path_file)

	# Pendiente (pdf has not only GRI 2016!!!)
	def GRI_GRI_OIL_GAS_COAL(self):
		wb = openpyxl.load_workbook(self.path_file)
		ws = wb[self.sheet]
		rows = ws.max_row

		# In depending on the option, it will mapp the data on the right columns on excel
		framework = input("'GRI 'oil & gas' or 'coal'???")
		for i in range(1, rows):
			if ws.cell(row=i, column=1).value == None:
				pass

			else:
				target_cell = ws.cell(row=i, column=2).value

		
				if target_cell != None and framework=='oil & gas':
					try:
						value_to_add = self.df.loc[self.df['GRI Code'] == target_cell]['SECTOR \nSTANDARD \nREF #'].values

						if len(value_to_add) > 0 and self.sheet== 'GRI 2016':
							ws.cell(row=i, column=18, value=' '.join(value_to_add))

						elif len(value_to_add) > 0 and self.sheet== 'GRI 2021':
							ws.cell(row=i, column=6, value='\n'.join(value_to_add))

					except:
						pass
				
				if target_cell != None and framework=='coal':
					try:
						value_to_add = self.df.loc[self.df['GRI Code'] == target_cell]['SECTOR \nSTANDARD \nREF #'].values

						if len(value_to_add) > 0 and self.sheet== 'GRI 2016':
							ws.cell(row=i, column=20, value='\n'.join(value_to_add))

						elif len(value_to_add) > 0 and self.sheet== 'GRI 2021':
							ws.cell(row=i, column=8, value='\n'.join(value_to_add))

					except:
						pass

		wb.save(self.path_file)

		return f'{self.sheet} sheet from Excel file have bee mapped'

	def mapGRI_HKEX22(self):
		wb = openpyxl.load_workbook(self.path_file)
		ws = wb[self.sheet]
		rows = ws.max_row

		for i in range(3, rows):
			if ws.cell(row=i, column=2).value == None:
				pass

			else:
				target_cell = ws.cell(row=i, column=2).value
		
				if target_cell != None:
					try:
						disclosure_to_add = self.df.loc[self.df['GRI_Code'] == target_cell]['HKEX ESG Reporting Guide'].values
						
						if len(disclosure_to_add) > 0:
							ws.cell(row=i, column=22, value='\n'.join(disclosure_to_add))
						
					except:
							pass

		wb.save(self.path_file)
		print(f"{self.sheet} sheet from Excel file have bee mapped with it's HKEX equivalent") 

	def mapHKEX22_GRI(self):
		wb = openpyxl.load_workbook(self.path_file)
		ws = wb[self.sheet]

		r=3
		for i in self.df["HKEX ESG Reporting Guide"]:
			if i != None:
				ws.cell(row=r, column=1, value=i)
				r+=1

		r=3
		for i in self.df['GRI Standards and Disclosures']:
			if i != None:
				ws.cell(row=r, column=2, value=i)
				r+=1

		r=3
		for i in self.df['GRI_Code']:
			if i != None:
				ws.cell(row=r, column=3, value=i)
				r+=1
		
		wb.save(self.path_file)
	
	def mapGRI16_BESI(self):
		wb = openpyxl.load_workbook(self.path_file)
		ws = wb[self.sheet]
		rows = ws.max_row

		for i in range(3, rows):
			if ws.cell(row=i, column=2).value == None:
				pass

			else:
				target_cell = ws.cell(row=i, column=2).value

				if target_cell != None:
					
					try:
						disclosure_to_add = self.df.loc[self.df['GRI_Disclosures_2016'] == target_cell]['SEBI - BRSR Framework'].values
						
						if len(disclosure_to_add) > 0:
							ws.cell(row=i, column=24, value='\n'.join(disclosure_to_add))
						
					except:
							pass

		wb.save(self.path_file)
		print(f"{self.sheet} sheet from Excel file have bee mapped with it's BESI BRSB equivalent") 

	def mapGRI21_BESI(self):
		wb = openpyxl.load_workbook(self.path_file)
		ws = wb[self.sheet]
		rows = ws.max_row

		for i in range(3, rows):
			if ws.cell(row=i, column=2).value == None:
				pass

			else:
				target_cell = ws.cell(row=i, column=2).value

				if target_cell != None:
					
					try:
						disclosure_to_add = self.df.loc[self.df['GRI_Disclosures_2021'] == target_cell]['SEBI - BRSR Framework'].values
						
						if len(disclosure_to_add) > 0:
							ws.cell(row=i, column=16, value='\n'.join(disclosure_to_add))
						
					except:
							pass

		wb.save(self.path_file)
		print(f"{self.sheet} sheet from Excel file have bee mapped with it's BESI BRSB equivalent") 

	def mapESGsRef(self):
		wb = openpyxl.load_workbook(self.path_file)
		ws = wb[self.sheet]
		rows = ws.max_row

		for i in range(self.first_row, rows):
			if ws.cell(row=i, column=self.excel_ref_column).value == None:
				pass

			else:
				target_cell = int(ws.cell(row=i, column=self.excel_ref_column).value[0:3])
		
				if target_cell != None:
					try:
						disclosure_to_add = self.df.loc[self.df[self.df_ref_column_str] == target_cell][self.df_ref_column_to_add].values
						
						
						if len(disclosure_to_add) > 0:
							ws.cell(row=i, column=self.excel_column_to_add, value='\n'.join(disclosure_to_add))
						
					except:
							pass

		wb.save(self.path_file)
		print(f"{self.sheet} sheet from Excel file have bee mapped with it's {self.ESG_to_add} equivalent") 


	def mapESGs(self):
		wb = openpyxl.load_workbook(self.path_file)
		ws = wb[self.sheet]

		
		for i in self.df.columns:
			r=self.first_row
			for c in self.df[i]:
				if c != None:
					ws.cell(row=r, column=self.df.columns.get_loc(i)+1, value=c)
					r+=1
			
		
			
		wb.save(self.path_file)




